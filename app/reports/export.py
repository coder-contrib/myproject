import io
import os
import logging
from typing import Optional
from datetime import datetime
from pathlib import Path

from app.reports.config import report_config

logger = logging.getLogger("reports.export")


class ReportExporter:
    """Exports reports to PDF and Excel formats."""

    def __init__(self):
        self.export_path = Path(report_config.export_path)
        self.export_path.mkdir(parents=True, exist_ok=True)

    def export_to_excel(
        self,
        data: dict,
        filename: Optional[str] = None,
        sheets: Optional[dict] = None,
    ) -> str:
        """Export report data to Excel (.xlsx) file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        if sheets:
            # Multiple sheets
            for i, (sheet_name, sheet_data) in enumerate(sheets.items()):
                if i == 0:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)
                self._write_sheet(ws, sheet_data)
        else:
            ws = wb.active
            ws.title = data.get("report_type", "Report")
            self._write_report_sheet(ws, data)

        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            report_type = data.get("report_type", "report")
            filename = f"{report_type}_{timestamp}.xlsx"

        filepath = self.export_path / filename
        wb.save(str(filepath))
        logger.info("Excel exported: %s", filepath)
        return str(filepath)

    def export_to_pdf(
        self,
        data: dict,
        filename: Optional[str] = None,
        template: Optional[str] = None,
    ) -> str:
        """Export report data to PDF file."""
        html_content = self._render_report_html(data, template)

        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            report_type = data.get("report_type", "report")
            filename = f"{report_type}_{timestamp}.pdf"

        filepath = self.export_path / filename

        if report_config.pdf_engine == "weasyprint":
            self._export_pdf_weasyprint(html_content, str(filepath))
        else:
            self._export_pdf_fallback(html_content, str(filepath))

        logger.info("PDF exported: %s", filepath)
        return str(filepath)

    def export_to_csv(
        self,
        rows: list[dict],
        filename: Optional[str] = None,
    ) -> str:
        """Export tabular data to CSV."""
        import csv

        if not rows:
            raise ValueError("No data to export")

        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.csv"

        filepath = self.export_path / filename
        columns = list(rows[0].keys())

        with open(filepath, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            for row in rows[:report_config.max_export_rows]:
                writer.writerow({k: str(v) for k, v in row.items()})

        logger.info("CSV exported: %s (%d rows)", filepath, len(rows))
        return str(filepath)

    def _write_sheet(self, ws, data: list[dict]):
        from openpyxl.styles import Font, PatternFill, Alignment

        if not data:
            return

        headers = list(data[0].keys())
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(data[:report_config.max_export_rows], 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                if isinstance(value, (list, dict)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            from openpyxl.utils import get_column_letter
            max_length = max(
                len(str(header)),
                max((len(str(row.get(header, ""))) for row in data[:100]), default=0)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    def _write_report_sheet(self, ws, data: dict):
        from openpyxl.styles import Font, Alignment

        # Title
        report_type = data.get("report_type", "Report").replace("_", " ").title()
        ws.cell(row=1, column=1, value=report_config.company_name).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=report_type).font = Font(bold=True, size=12)

        if "period" in data:
            period = data["period"]
            ws.cell(row=3, column=1, value=f"Period: {period.get('start', '')} to {period.get('end', '')}")
        elif "as_of_date" in data:
            ws.cell(row=3, column=1, value=f"As of: {data['as_of_date']}")

        ws.cell(row=4, column=1, value=f"Generated: {data.get('generated_at', '')}")

        row = 6
        for key, value in data.items():
            if key in ("report_type", "period", "as_of_date", "generated_at"):
                continue

            if isinstance(value, dict):
                ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = Font(bold=True)
                row += 1
                for sub_key, sub_val in value.items():
                    ws.cell(row=row, column=1, value=f"  {sub_key.replace('_', ' ').title()}")
                    ws.cell(row=row, column=2, value=sub_val if not isinstance(sub_val, (list, dict)) else str(sub_val))
                    row += 1
                row += 1
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = Font(bold=True)
                row += 1
                self._write_table_at_row(ws, value, row)
                row += len(value) + 2
            else:
                ws.cell(row=row, column=1, value=key.replace("_", " ").title())
                ws.cell(row=row, column=2, value=value if not isinstance(value, (list, dict)) else str(value))
                row += 1

    def _write_table_at_row(self, ws, data: list[dict], start_row: int):
        from openpyxl.styles import Font

        if not data:
            return

        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col_idx, value=header.replace("_", " ").title()).font = Font(bold=True)

        for row_idx, row_data in enumerate(data[:50], start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                if isinstance(value, (list, dict)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _render_report_html(self, data: dict, template: Optional[str] = None) -> str:
        report_type = data.get("report_type", "Report").replace("_", " ").title()
        currency = report_config.currency_symbol

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: 'Helvetica Neue', Arial, sans-serif; margin: 40px; color: #333; }}
                h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
                h2 {{ color: #34495e; margin-top: 30px; }}
                table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
                th {{ background-color: #3498db; color: white; padding: 10px; text-align: left; }}
                td {{ padding: 8px 10px; border-bottom: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f8f9fa; }}
                .metric {{ display: inline-block; margin: 10px 20px 10px 0; padding: 15px;
                          background: #f8f9fa; border-radius: 8px; min-width: 150px; }}
                .metric-label {{ font-size: 0.85em; color: #666; }}
                .metric-value {{ font-size: 1.5em; font-weight: bold; color: #2c3e50; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .footer {{ margin-top: 40px; text-align: center; font-size: 0.8em; color: #999; }}
                .positive {{ color: #27ae60; }}
                .negative {{ color: #e74c3c; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{report_config.company_name}</h1>
                <h2>{report_type}</h2>
        """

        if "period" in data:
            html += f"<p>Period: {data['period'].get('start', '')} to {data['period'].get('end', '')}</p>"
        elif "as_of_date" in data:
            html += f"<p>As of: {data['as_of_date']}</p>"

        html += "</div>"

        for key, value in data.items():
            if key in ("report_type", "period", "as_of_date", "generated_at"):
                continue

            section_title = key.replace("_", " ").title()

            if isinstance(value, dict):
                html += f"<h2>{section_title}</h2><div>"
                for sub_key, sub_val in value.items():
                    if isinstance(sub_val, (int, float)):
                        formatted = f"{currency}{sub_val:,.2f}" if "revenue" in sub_key or "total" in sub_key or "amount" in sub_key or "income" in sub_key or "cost" in sub_key or "profit" in sub_key else f"{sub_val:,.2f}"
                        html += f'<div class="metric"><div class="metric-label">{sub_key.replace("_", " ").title()}</div><div class="metric-value">{formatted}</div></div>'
                    elif not isinstance(sub_val, (list, dict)):
                        html += f'<div class="metric"><div class="metric-label">{sub_key.replace("_", " ").title()}</div><div class="metric-value">{sub_val}</div></div>'
                html += "</div>"

            elif isinstance(value, list) and value and isinstance(value[0], dict):
                html += f"<h2>{section_title}</h2><table><thead><tr>"
                headers = list(value[0].keys())
                for h in headers:
                    html += f"<th>{h.replace('_', ' ').title()}</th>"
                html += "</tr></thead><tbody>"
                for row in value[:50]:
                    html += "<tr>"
                    for h in headers:
                        cell_val = row.get(h, "")
                        if isinstance(cell_val, float):
                            cell_val = f"{cell_val:,.2f}"
                        html += f"<td>{cell_val}</td>"
                    html += "</tr>"
                html += "</tbody></table>"

            elif isinstance(value, (int, float)):
                formatted = f"{currency}{value:,.2f}" if any(kw in key for kw in ["revenue", "income", "cost", "profit", "total", "flow"]) else f"{value:,.2f}"
                css_class = "positive" if value >= 0 else "negative"
                html += f'<div class="metric"><div class="metric-label">{section_title}</div><div class="metric-value {css_class}">{formatted}</div></div>'

        html += f"""
            <div class="footer">
                <p>Generated: {data.get('generated_at', datetime.utcnow().isoformat())} | {report_config.company_name}</p>
            </div>
        </body>
        </html>
        """

        return html

    def _export_pdf_weasyprint(self, html: str, filepath: str):
        try:
            from weasyprint import HTML
            HTML(string=html).write_pdf(filepath)
        except ImportError:
            logger.warning("weasyprint not installed, using fallback")
            self._export_pdf_fallback(html, filepath)

    def _export_pdf_fallback(self, html: str, filepath: str):
        """Save HTML as-is when PDF engine unavailable."""
        html_path = filepath.replace(".pdf", ".html")
        with open(html_path, "w") as f:
            f.write(html)
        logger.info("PDF engine unavailable, saved as HTML: %s", html_path)


report_exporter = ReportExporter()
